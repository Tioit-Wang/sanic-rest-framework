"""
@Author：WangYuXiang
@E-mile：Hill@3io.cc
@CreateTime：2021/6/2 14:58
@DependencyLibrary：无
@MainFunction：无
@FileDoc： 
    export_import_view.py
    文件说明
@ChangeHistory:
    datetime action why
    example:
    2021/6/2 14:58 change 'Fix bug'
        
"""
import io
import time

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from sanic.response import raw

from srf.exceptions import APIException
from srf.mixins import ListModelMixin, CreateModelMixin
from srf.status import HttpStatus
from srf.viewsets import GenericViewSet


class ExportImportView(ListModelMixin, CreateModelMixin, GenericViewSet):
    """导入导出基类"""
    export_header_title = {}
    import_header_title = {}
    upload_file_key = 'import_file'

    async def list(self, request, *args, **kwargs):
        queryset = await self.get_queryset()
        page = await self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
        else:
            serializer = self.get_serializer(queryset, many=True)

        virtual_workbook = self.write_data_to_excel(await serializer.data)

        return raw(save_virtual_workbook(virtual_workbook), headers={
            'Content-Disposition': 'attachment;filename={}'.format(self.get_export_file_name())
        })

    async def create(self, request, *args, **kwargs):
        workbook: Workbook = self.load_workbook_by_request()
        worksheet = workbook.worksheets[0]
        rows = worksheet.max_row  # 获取表的最大行数
        columns = worksheet.max_column  # 获取表的最大列数

        # 检查导入数据是否必备
        for column in range(1, columns + 1):
            chinese_key = worksheet.cell(1, column).value
            if chinese_key not in self.import_header_title:
                return self.error_json_response('存在意外的数据列%s' % chinese_key)

        for row in range(2, rows + 1):
            row_record = {}
            for column in range(1, columns + 1):
                chinese_key = worksheet.cell(1, column).value
                letter_key = self.import_header_title[chinese_key]
                row_record[letter_key] = worksheet.cell(row, column).value
            serializer = self.serializer_class(data=row_record)
            await serializer.is_valid(raise_exception=True)
            await self.perform_create(serializer)
        return self.success_json_response(msg='导入成功', http_status=HttpStatus.HTTP_201_CREATED)

    def load_workbook_by_request(self):
        """从请求中加载Excel"""
        upload_file = self.request.files.get(self.upload_file_key)
        if upload_file:
            raise APIException(message='必须上传导入文件 file_name.xlsx', http_status=HttpStatus.HTTP_200_OK)
        stream = upload_file.body
        workbook: Workbook = load_workbook(io.BytesIO(stream))
        return workbook

    def get_export_file_name(self):
        """导出名"""
        return 'export_{}.xlsx'.format(int(time.time()))

    def create_excel(self):
        """导出执行"""
        virtual_workbook = Workbook()
        virtual_workbook.create_sheet('export', 0)
        virtual_workbook.remove_sheet(virtual_workbook.worksheets[1])
        return virtual_workbook

    def write_data_to_excel(self, data):
        """为excel写入数据"""
        virtual_workbook = self.create_excel()
        sheet = virtual_workbook.worksheets[0]
        for x_index, key in enumerate(self.export_header_title):
            x = x_index + 1
            title = self.export_header_title[key]
            sheet[self.get_index(x, 1)] = title
            for y_index, row in enumerate(data):
                y = y_index + 2
                sheet[self.get_index(x, y)] = row[key]
        return virtual_workbook

    def get_index(self, x, y):
        """得到字母+数字坐标"""
        return '{}{}'.format(self.get_letter_head(x), y)

    def get_letter_head(self, x):
        """数字转字母"""
        start_num = ord('A')
        if x <= 26:
            return chr(start_num + x - 1)
        else:
            return self.get_letter_head((x - 1) // 26) + chr(start_num + (x - 1) % 26)
